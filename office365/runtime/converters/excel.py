"""Excel (.xlsx) exporter/importer via openpyxl — the optional-dependency boundary."""

from __future__ import annotations

from os import PathLike
from typing import Any, Dict, List, Union, cast


def require_openpyxl():
    """Import openpyxl or raise an actionable error.

    Raises:
        ImportError: When openpyxl is not installed, pointing at the extra.
    """
    try:
        import openpyxl  # type: ignore[import-not-found]
    except ImportError:
        raise ImportError("pip install office365-rest-python-client[excel]") from None
    return openpyxl


def write_excel(records: List[Dict[str, Any]], path: Union[str, PathLike]) -> None:
    """Write records to an Excel worksheet (header row + one row per record)."""
    openpyxl = require_openpyxl()
    workbook = openpyxl.Workbook()
    worksheet = cast(Any, workbook.active)
    if records:
        columns = list(dict.fromkeys(key for record in records for key in record))
        worksheet.append(columns)
        for record in records:
            worksheet.append([record.get(column) for column in columns])
    workbook.save(path)


def read_excel(path: Union[str, PathLike]) -> List[Dict[str, Any]]:
    """Read the first worksheet into records (header row -> keys)."""
    openpyxl = require_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=True)
    worksheet = cast(Any, workbook.active)
    rows = worksheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if headers is None:
        return []
    return [dict(zip(headers, row)) for row in rows if any(value is not None for value in row)]
